from openpyxl import load_workbook
import csv

# load the workbook and select the active sheet
wb = load_workbook('CourseDefinitions UC Interface export Kash 19.4.23.xlsx')
sheet = wb.active


def strip_if_str(val):
    if isinstance(val, str):
        return val.strip()
    else:
        return val


# write the rows to a CSV
# name of csv file
filename = "rows.csv"

# iterate through the rows
count = 0
with open(filename, 'w', newline='') as csvfile:
    csvwriter = csv.writer(csvfile)
    for row in sheet.iter_rows(values_only=True):
        csvwriter.writerow([strip_if_str(x) for x in row])
        count += 1
        if count > 100:
            break


