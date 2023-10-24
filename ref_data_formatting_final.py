import pandas as pd
import openpyxl

from openpyxl.styles import PatternFill, Font
from ortools.sat.python import cp_model


def distribute_objects(objects):
    model = cp_model.CpModel()

    num_objects = len(objects)
    total_size = sum(objects.values())
    object_values = list(objects.values())
    num_containers = 4

    # Create decision variables.
    # x[i][j] is 1 if object i is in container j; 0 otherwise.
    x = [[model.NewIntVar(0, 1, f'x_{i}_{j}') for j in range(num_containers)] for i in range(num_objects)]

    # Each object should be in exactly one container.
    for i in range(num_objects):
        model.Add(sum(x[i][j] for j in range(num_containers)) == 1)

    # Define the total size in each container.
    container_sizes = [model.NewIntVar(0, total_size, f'size_{j}') for j in range(num_containers)]
    for j in range(num_containers):
        model.Add(sum(x[i][j] * object_values[i] for i in range(num_objects)) == container_sizes[j])

    # Define the objective: minimize the difference between the maximum and minimum total sizes.
    max_size = model.NewIntVar(0, sum(objects), 'max_size')
    min_size = model.NewIntVar(0, sum(objects), 'min_size')
    model.AddMaxEquality(max_size, container_sizes)
    model.AddMinEquality(min_size, container_sizes)
    model.Minimize(max_size - min_size)

    # Solve the problem.
    solver = cp_model.CpSolver()
    status = solver.Solve(model)

    if status == cp_model.OPTIMAL:
        assignments = [[solver.Value(x[i][j]) for j in range(num_containers)] for i in range(num_objects)]
        return assignments
    else:
        return None


# Load the data
data_sheet = pd.read_excel('config_template.xlsx', sheet_name='data')

# Remove any rows where the value is empty
data_sheet = data_sheet[data_sheet['Value'].notna()]

# Group the data by TableName and ColumnName and count the number of rows in each group
# (adding a buffer for the header rows)
group_sizes = data_sheet.groupby(['TableName', 'ColumnName']).size() + 4

# Build a dictionary of group_size keys and values
objects = {}
for idx, group_size in enumerate(group_sizes):
    objects[idx] = group_size

assignments = distribute_objects(objects)

# Print the container assigments
if assignments:
    for i, obj_assign in enumerate(assignments):
        container = obj_assign.index(1)
        table_column_key = group_sizes.index[i]
        print(f'Object  {table_column_key} is in container {container + 1}')
else:
    print("No solution found!")

# Print the size of each container
container_sizes = [sum([group_sizes[i] for i, obj_assign in enumerate(assignments) if obj_assign.index(1) == j]) for j in range(4)]
print(f'Container sizes: {container_sizes}')

# Add each object key/value pair to the appropriate container
bins = [[] for _ in range(4)]
for i, obj_assign in enumerate(assignments):
    container = obj_assign.index(1)
    bins[container].append(group_sizes.index[i])

# Create output Excel workbook
wb = openpyxl.Workbook()
del wb['Sheet']

header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
bold_font = Font(bold=True)

for idx, bin in enumerate(bins):
    ws = wb.create_sheet(title=f"Group {idx + 1}")
    current_row = 1
    for group in bin:
        group_data = data_sheet[data_sheet['TableName'] == group[0]][data_sheet['ColumnName'] == group[1]]

        # Table header rows
        ws.cell(row=current_row, column=1, value=f"{group[0]}.{group[1]}").font = bold_font
        ws.cell(row=current_row, column=1).fill = header_fill
        current_row += 2
        ws.cell(row=current_row, column=1, value='Key').font = bold_font
        ws.cell(row=current_row, column=2, value='Value').font = bold_font
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=2).fill = header_fill
        current_row += 1

        for _, row in group_data.iterrows():
            ws.cell(row=current_row, column=1, value=row['RowNum'])
            ws.cell(row=current_row, column=2, value=row['Value'])
            current_row += 1

        current_row += 1

wb.save('grouped_config_template_optimal.xlsx')

print("Optimal grouping saved to 'grouped_config_template_optimal.xlsx'")


