import pandas as pd
from ortools.algorithms.python import knapsack_solver
import openpyxl
from openpyxl.styles import PatternFill, Font

# Load the data
data_sheet = pd.read_excel('config_template.xlsx', sheet_name='data')
group_sizes = data_sheet.groupby(['TableName', 'ColumnName']).size() + 3

# Initialize the solver
solver = knapsack_solver.KnapsackSolver(
    knapsack_solver.KNAPSACK_MULTIDIMENSION_BRANCH_AND_BOUND_SOLVER, 'BinPacking')

weights = group_sizes.values.tolist()
capacities = [max(weights)] * 4

solver.init(weights, [weights], capacities)
solver.solve()

bins = [[] for _ in range(4)]
for i in range(len(weights)):
    bin_idx = solver.BestSolutionContains(i)
    bins[bin_idx].append(group_sizes.index[i])

# Create output Excel workbook
wb = openpyxl.Workbook()
del wb['Sheet']

header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
bold_font = Font(bold=True)

for idx, bin in enumerate(bins):
    ws = wb.create_sheet(title=f"Group {idx + 1}")
    current_row = 1
    for group in bin:
        group_data = data_sheet[data_sheet['TableName'] == group[0]][data_sheet['ColumnName'] == group[1]]

        ws.cell(row=current_row, column=1, value=f"{group[0]}.{group[1]}").font = bold_font
        ws.cell(row=current_row, column=1).fill = header_fill
        current_row += 1

        for _, row in group_data.iterrows():
            ws.cell(row=current_row, column=1, value=row['Value'])
            ws.cell(row=current_row, column=2, value=row['RowNum'])
            current_row += 1

        current_row += 1

wb.save('grouped_config_template_optimal.xlsx')

print("Optimal grouping saved to 'grouped_config_template_optimal.xlsx'")
