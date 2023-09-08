from openpyxl import Workbook

relationships = [
    {'child_id': 1, 'child_name': 'Child 1', 'parent_id': None},
    {'child_id': 2, 'child_name': 'Child 2', 'parent_id': 1},
    {'child_id': 3, 'child_name': 'Child 3', 'parent_id': 1},
    {'child_id': 4, 'child_name': 'Child 4', 'parent_id': 3},
    {'child_id': 5, 'child_name': 'Child 5', 'parent_id': 3},
]


# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Write header
ws.append(['Child ID', 'Child Name', 'Parent ID'])


# Function to write rows and create groups
def write_rows_and_group(parent_id, level):
    row_start = ws.max_row + 1
    for relationship in relationships:
        if relationship['parent_id'] == parent_id:
            ws.append([relationship['child_id'], relationship['child_name'], relationship['parent_id']])
            write_rows_and_group(relationship['child_id'], level + 1)
    row_end = ws.max_row
    if row_start < row_end:
        ws.row_dimensions.group(start=row_start + 1, end=row_end, outline_level=level, hidden=True)


# Populate the worksheet and create groups
write_rows_and_group(None, 0)

# Save the workbook
wb.save('grouped_data.xlsx')
